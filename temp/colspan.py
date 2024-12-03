import pandas as pd
from openpyxl import load_workbook
from typing import Dict, Tuple

class MergeChecker:
    def __init__(self, excel_file: str):
        # DataFrame 로드
        self.df = pd.read_excel(excel_file)
        
        # openpyxl로 병합 정보 로드
        wb = load_workbook(excel_file)
        self.sheet = wb.active
        
        # 병합 정보 캐시 생성
        self.merge_cache = self._create_merge_cache()
    
    def _create_merge_cache(self) -> Dict[Tuple[int, int], int]:
        """병합 정보를 미리 저장하는 캐시 생성"""
        cache = {}
        for merged_range in self.sheet.merged_cells.ranges:
            # 병합 범위의 모든 셀에 대해 병합 너비 저장
            width = merged_range.max_col - merged_range.min_col + 1
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    cache[(row, col)] = width
        return cache
    
    def get_merge_width(self, row_idx: int, col_idx: int) -> int:
        """특정 셀의 가로 병합 수 반환"""
        # DataFrame 인덱스를 Excel 행 번호로 변환 (헤더 고려)
        excel_row = row_idx + 2  # 헤더가 있는 경우
        excel_col = col_idx + 1  # pandas는 0부터, excel은 1부터 시작
        
        # 캐시에서 병합 정보 확인
        return self.merge_cache.get((excel_row, excel_col), 1)

# 사용 예시
def main():
    excel_file = "your_excel_file.xlsx"
    checker = MergeChecker(excel_file)
    
    # DataFrame 순회하면서 병합 정보 확인
    for idx, row in checker.df.iterrows():
        print(f"\n=== 행 {idx + 2} ===")  # Excel 행 번호로 표시
        
        # 각 열의 병합 수 확인
        for col_idx in range(len(row)):
            merge_width = checker.get_merge_width(idx, col_idx)
            value = row[col_idx]
            print(f"열 {col_idx + 1}: 값 = {value}, 가로 병합 = {merge_width}칸")

# 간단한 사용 예시
if __name__ == "__main__":
    checker = MergeChecker("your_excel_file.xlsx")
    
    # 특정 row의 특정 열 병합 수 확인
    for idx, row in checker.df.iterrows():
        # row[3]의 병합 수 확인 (4번째 열)
        merge_width = checker.get_merge_width(idx, 3)
        value = row[3]
        print(f"행 {idx + 2}, 열 4: 값 = {value}, 가로 병합 = {merge_width}칸")


# 사용사레  1
# # 초기화
# checker = MergeChecker("your_excel_file.xlsx")

# # DataFrame 순회하면서 row[3] 확인
# for idx, row in checker.df.iterrows():
#     merge_width = checker.get_merge_width(idx, 3)  # 4번째 열 확인
#     print(f"행 {idx + 2}의 4번째 열 병합 수: {merge_width}")

# 사용사레  1
# checker = MergeChecker("your_excel_file.xlsx")

# # 특정 행의 모든 열 병합 정보 확인
# row_idx = 0  # 첫 번째 데이터 행
# row = checker.df.iloc[row_idx]
# for col_idx, value in enumerate(row):
#     merge_width = checker.get_merge_width(row_idx, col_idx)
#     if merge_width > 1:
#         print(f"열 {col_idx + 1}은 {merge_width}칸 병합되어 있습니다. 값: {value}")